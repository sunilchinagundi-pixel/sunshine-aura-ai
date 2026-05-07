from fastapi import FastAPI, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, EmailStr
from enum import Enum
from typing import List, Optional
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from sqlalchemy.orm import Session
from database import engine, SessionLocal, Base, User, CallbackRequest, get_db

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="Sunshine Aura AI Backend",
    description="Training, Consultancy, Jobs & Learning Platform",
    version="0.1.0"
)

# Enable CORS for React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="dist"), name="static")

# Models
class Role(str, Enum):
    free = "free"
    pro = "pro"

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    mobile: str
    role: Role = Role.free

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    id: int
    name: str
    email: str
    mobile: str
    role: str
    created_at: datetime

    class Config:
        from_attributes = True

class CallbackRequestCreate(BaseModel):
    name: str
    email: EmailStr
    mobile: str
    course: str
    location: Optional[str] = None
    message: Optional[str] = None

class Module(BaseModel):
    number: int
    title: str
    topics: List[str]

class Project(BaseModel):
    name: str
    description: str

class Training(BaseModel):
    id: int
    title: str
    description: str
    category: str
    is_free: bool
    tools: Optional[List[str]] = None
    internship: Optional[bool] = False
    certification: Optional[bool] = False
    duration: Optional[str] = None
    modules: Optional[List[Module]] = None
    projects: Optional[List[Project]] = None
    job_roles: Optional[List[str]] = None
    salary_range: Optional[str] = None
    schedule: Optional[str] = None

class JobPost(BaseModel):
    id: int
    title: str
    company: str
    location: str
    job_type: str
    description: str

class ConsultingOffer(BaseModel):
    id: int
    title: str
    details: str

# Static data
trainings = [
    Training(
        id=1,
        title="Python Fundamentals",
        description="Master Python basics for business applications, scripting, and automation.",
        category="Python",
        is_free=True,
        tools=["Python 3.11", "Jupyter Notebook", "VS Code", "Git"],
        internship=False,
        certification=True,
        duration="4 weeks",
        schedule="Weekdays (6-8 PM)",
        salary_range="₹5-15 LPA",
        job_roles=["Python Developer", "Automation Engineer", "Data Engineer (Entry)"],
        modules=[
            Module(number=1, title="Python Basics & Setup", topics=["Variables, Data Types", "Loops, Functions", "OOP Concepts"]),
            Module(number=2, title="Exception Handling & File Operations", topics=["Try-Except Blocks", "File I/O", "Working with Databases (SQL)"]),
            Module(number=3, title="Libraries & Frameworks Intro", topics=["Pandas Basics", "NumPy Arrays", "REST API Basics"]),
        ],
        projects=[
            Project(name="Automation Script", description="Create a script to automate daily tasks"),
            Project(name="Data Processing Tool", description="Build a tool to process and analyze CSV files"),
        ]
    ),
    Training(
        id=2,
        title="AI + Python Web Applications",
        description="Build modern AI-powered web apps using FastAPI, integrate LLMs, RAG, and deploy to cloud. Weekend program!",
        category="AI",
        is_free=True,
        tools=["Python", "FastAPI", "TensorFlow", "OpenAI API", "Docker", "AWS"],
        internship=True,
        certification=True,
        duration="6 weeks",
        schedule="Friday • Saturday • Sunday (10 AM - 5 PM)",
        salary_range="₹8-25 LPA",
        job_roles=["AI Developer", "Backend Engineer", "Full-Stack Developer", "ML Engineer"],
        modules=[
            Module(number=1, title="Python & Backend Fundamentals", topics=["Variables, Loops, Functions", "OOP Concepts", "Exception Handling", "REST API Basics", "Working with Databases (SQL)"]),
            Module(number=2, title="Web Development with FastAPI", topics=["Building APIs with FastAPI", "Request & Response Handling", "Data Validation (Pydantic)", "API Testing with Swagger", "Mini Backend Project"]),
            Module(number=3, title="AI Fundamentals", topics=["Introduction to AI & NLP", "Text Processing Basics", "Transformers Overview", "Using Pre-trained Models (HuggingFace)"]),
            Module(number=4, title="LLM & AI Integration", topics=["Prompt Engineering Basics", "Working with LLMs", "OpenAI & Open Source Models", "Building AI-powered Features", "Chatbot Development"]),
            Module(number=5, title="RAG (Retrieval-Augmented Generation)", topics=["Concept of RAG", "Document Processing", "Vector Databases (FAISS, ChromaDB)", "Build AI Document Q&A System"]),
            Module(number=6, title="Cloud & Deployment", topics=["AWS Basics (EC2, S3)", "Deploy FastAPI Applications", "Docker Basics", "Hosting AI Applications"]),
            Module(number=7, title="DevOps Basics", topics=["Git & GitHub", "CI/CD Basics", "Automated Deployment", "Monitoring & Logs"]),
            Module(number=8, title="Projects (Hands-On)", topics=["AI Chatbot", "AI Document Assistant", "AI Interview System", "Resume Analyzer", "Sentiment Analysis API"]),
            Module(number=9, title="Teach & Earn (Guidance)", topics=["How to Create Content", "Build Online Presence", "Freelancing Basics", "Mentorship Tips"]),
        ],
        projects=[
            Project(name="AI Chatbot", description="Build an intelligent chatbot using OpenAI API"),
            Project(name="AI Document Assistant", description="Create a system to process and query documents with AI"),
            Project(name="AI Interview System", description="Build an AI-powered interview preparation tool"),
            Project(name="Resume Analyzer", description="Analyze and score resumes using NLP"),
            Project(name="Sentiment Analysis API", description="Deploy a sentiment analysis API to production"),
        ]
    ),
    Training(
        id=3,
        title="Advanced AI Architecture (Pro)",
        description="Production-grade AI systems: LLMs, Neural Networks, MLOps, Cloud Deployment.",
        category="AI",
        is_free=False,
        tools=["PyTorch", "TensorFlow", "Hugging Face", "AWS SageMaker", "Kubernetes", "MLflow"],
        internship=True,
        certification=True,
        duration="8 weeks",
        salary_range="₹15-40 LPA",
        job_roles=["Senior AI Engineer", "ML Architect", "Research Engineer"],
        modules=[
            Module(number=1, title="Advanced Deep Learning", topics=["Neural Network Architectures", "CNNs & RNNs", "Attention Mechanisms", "Transformers"]),
            Module(number=2, title="LLM Fine-Tuning", topics=["Model Selection & Setup", "Fine-tuning Techniques", "LoRA & QLoRA", "Optimization Strategies"]),
            Module(number=3, title="MLOps & Production", topics=["Model Versioning", "Pipeline Orchestration", "Model Monitoring", "A/B Testing"]),
        ]
    ),
    Training(
        id=4,
        title="AI + Machine Learning Fundamentals",
        description="Learn supervised & unsupervised learning, classification, regression, and real-world ML projects.",
        category="AI",
        is_free=True,
        tools=["Scikit-learn", "Pandas", "NumPy", "Matplotlib", "Jupyter"],
        internship=False,
        certification=True,
        duration="5 weeks",
        salary_range="₹6-20 LPA",
        job_roles=["ML Engineer", "Data Scientist", "Analytics Engineer"],
        modules=[
            Module(number=1, title="ML Fundamentals", topics=["Supervised Learning", "Unsupervised Learning", "Model Evaluation", "Cross-Validation"]),
            Module(number=2, title="Algorithms & Techniques", topics=["Linear & Logistic Regression", "Decision Trees", "K-Means Clustering", "Dimensionality Reduction"]),
            Module(number=3, title="Real-World Projects", topics=["Feature Engineering", "Data Preprocessing", "Building ML Pipelines"]),
        ]
    ),
    Training(
        id=5,
        title="Full-Stack AI Development (Internship + Certification)",
        description="Complete pathway: Python → Web Dev → AI/ML → Deploy to Cloud. Build real projects with mentorship.",
        category="AI",
        is_free=False,
        tools=["Python", "React", "FastAPI", "TensorFlow", "AWS", "Docker", "Git"],
        internship=True,
        certification=True,
        duration="12 weeks",
        salary_range="₹12-35 LPA",
        job_roles=["Full-Stack AI Developer", "AI Engineer", "Tech Lead"],
        modules=[
            Module(number=1, title="Python Mastery", topics=["Core Python Concepts", "OOP & Design Patterns", "Testing & Debugging"]),
            Module(number=2, title="Frontend (React)", topics=["React Basics", "State Management", "API Integration"]),
            Module(number=3, title="Backend (FastAPI)", topics=["API Development", "Database Design", "Authentication"]),
            Module(number=4, title="AI/ML Integration", topics=["ML Model Development", "Integration with Backend", "Real-time Predictions"]),
            Module(number=5, title="Cloud Deployment", topics=["AWS Services", "Containerization", "Scaling Applications"]),
        ]
    ),
    Training(
        id=6,
        title="Generative AI & LLM Engineering",
        description="ChatGPT, Llama, BERT integration. Fine-tuning models, prompt engineering, RAG systems.",
        category="AI",
        is_free=False,
        tools=["Hugging Face", "OpenAI API", "LangChain", "Llama", "Prompt Engineering"],
        internship=True,
        certification=True,
        duration="6 weeks",
        salary_range="₹15-40 LPA",
        job_roles=["GenAI Engineer", "Prompt Engineer", "LLM Specialist"],
        modules=[
            Module(number=1, title="Generative AI Fundamentals", topics=["GANs & Diffusion Models", "Transformer Basics", "Attention Mechanisms"]),
            Module(number=2, title="LLM & Prompt Engineering", topics=["Working with OpenAI", "Prompt Optimization", "Fine-tuning LLMs"]),
            Module(number=3, title="RAG & Advanced Techniques", topics=["Retrieval-Augmented Generation", "Vector Stores", "Building Intelligent Systems"]),
        ]
    ),
]
jobs = [
    JobPost(id=1, title="Junior Python Trainer", company="Sunshine Aura AI", location="Remote", job_type="Full-time", description="Teach Python and AI to corporate teams."),
    JobPost(id=2, title="AI Consultant", company="Aura Consulting", location="Hybrid", job_type="Contract", description="Design AI solutions for enterprise clients."),
]
offers = [
    ConsultingOffer(id=1, title="AI Strategy Consulting", details="1-on-1 sessions to plan your AI adoption."),
    ConsultingOffer(id=2, title="Career Guidance", details="Personalized career coaching in AI and tech."),
]

# Admin token for protected routes
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "sunshine-admin-secret-2026")

def verify_admin(authorization: Optional[str] = Header(None)):
    if not authorization or authorization != f"Bearer {ADMIN_TOKEN}":
        raise HTTPException(status_code=401, detail="Unauthorized")
    return True

# Routes
@app.get("/")
def root():
    dist_path = os.path.join(os.path.dirname(__file__), "dist")
    index_path = os.path.join(dist_path, "index.html")
    if os.path.isfile(index_path):
        return FileResponse(index_path)
    return {"message": "Welcome to Sunshine Aura AI", "version": "0.1.0"}

@app.post("/api/register")
def register(user: UserCreate, db: Session = Depends(get_db)):
    # Check if email already exists
    existing_user = db.query(User).filter(User.email == user.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered.")
    
    # Create new user
    new_user = User(
        name=user.name,
        email=user.email,
        password=user.password,
        mobile=user.mobile,
        role=user.role.value
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return {
        "message": "Registration successful",
        "user": UserResponse.from_orm(new_user)
    }

@app.post("/api/login")
def login(credentials: UserLogin, db: Session = Depends(get_db)):
    user = db.query(User).filter(
        User.email == credentials.email,
        User.password == credentials.password
    ).first()
    
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials.")
    
    return {
        "message": "Login successful",
        "user": UserResponse.from_orm(user)
    }

@app.get("/api/trainings")
def get_trainings(q: Optional[str] = None):
    if q:
        query_lower = q.strip().lower()
        return [t for t in trainings if query_lower in t.title.lower() or query_lower in t.description.lower() or query_lower in t.category.lower()]
    return trainings

@app.get("/api/trainings/{course_id}")
def get_course_detail(course_id: int):
    for course in trainings:
        if course.id == course_id:
            return course
    raise HTTPException(status_code=404, detail="Course not found")

@app.post("/api/callback-request")
def create_callback_request(request: CallbackRequestCreate, db: Session = Depends(get_db)):
    callback = CallbackRequest(
        name=request.name,
        email=request.email,
        mobile=request.mobile,
        course=request.course,
        location=request.location,
        message=request.message,
    )
    db.add(callback)
    db.commit()
    db.refresh(callback)
    return {
        "message": "Callback request received. Our team will contact you soon.",
        "request_id": callback.id
    }

@app.get("/api/jobs", response_model=List[JobPost])
def get_jobs():
    return jobs

@app.get("/api/consulting", response_model=List[ConsultingOffer])
def get_consulting():
    return offers

@app.get("/api/pricing")
def get_pricing():
    return {
        "free": {
            "price": 0,
            "features": ["Basic trainings", "Job listings"]
        },
        "pro": {
            "price": 12999,
            "features": ["All trainings", "Consulting sessions", "Job placement support", "Premium community"]
        }
    }

@app.get("/api/users", response_model=List[UserResponse])
def get_all_users(db: Session = Depends(get_db)):
    users_list = db.query(User).all()
    return [UserResponse.from_orm(u) for u in users_list]

@app.get("/api/admin/export-users")
def export_users_excel(authorization: Optional[str] = Header(None), db: Session = Depends(get_db)):
    # Verify admin token
    if not authorization or authorization != f"Bearer {ADMIN_TOKEN}":
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Fetch all users
    users_list = db.query(User).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Registered Users"
    
    # Header row
    headers = ["ID", "Name", "Email", "Mobile", "Role", "Registered Date"]
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data rows
    for user in users_list:
        ws.append([
            user.id,
            user.name,
            user.email,
            user.mobile,
            user.role,
            user.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sunshine_aura_users.xlsx"}
    )

@app.get("/api/admin/callback-requests")
def get_callback_requests(authorization: Optional[str] = Header(None), db: Session = Depends(get_db)):
    if not authorization or authorization != f"Bearer {ADMIN_TOKEN}":
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    callbacks = db.query(CallbackRequest).order_by(CallbackRequest.created_at.desc()).all()
    return [
        {
            "id": c.id,
            "name": c.name,
            "email": c.email,
            "mobile": c.mobile,
            "course": c.course,
            "location": c.location,
            "message": c.message,
            "created_at": c.created_at
        }
        for c in callbacks
    ]

@app.get("/api/admin/export-callback-requests")
def export_callback_requests_excel(authorization: Optional[str] = Header(None), db: Session = Depends(get_db)):
    if not authorization or authorization != f"Bearer {ADMIN_TOKEN}":
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    callbacks = db.query(CallbackRequest).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Callback Requests"
    
    # Header row
    headers = ["ID", "Name", "Email", "Mobile", "Course", "Location", "Message", "Requested Date"]
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data rows
    for callback in callbacks:
        ws.append([
            callback.id,
            callback.name,
            callback.email,
            callback.mobile,
            callback.course,
            callback.location or "",
            callback.message or "",
            callback.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 20
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sunshine_aura_callback_requests.xlsx"}
    )

# Serve React app
@app.get("/{full_path:path}")
async def serve_react_app(full_path: str):
    if full_path.startswith("api/"):
        return {"error": "API route not found"}
    dist_path = os.path.join(os.path.dirname(__file__), "dist")
    file_path = os.path.join(dist_path, full_path)
    if os.path.isfile(file_path):
        return FileResponse(file_path)
    # Serve index.html for SPA routing
    index_path = os.path.join(dist_path, "index.html")
    if os.path.isfile(index_path):
        return FileResponse(index_path)
    return {"error": "File not found"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
